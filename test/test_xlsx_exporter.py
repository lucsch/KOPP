from datetime import date, datetime, timedelta

from openpyxl import load_workbook

from kopp.xlsx_exporter import XlsxExporter


def test_export_writes_headers_and_rows_to_xlsx(tmp_path):
    filename = tmp_path / "records.xlsx"

    XlsxExporter.export(
        filename,
        [
            {
                "day": "lundi",
                "date": date(2026, 1, 1),
                "hr": 90,
                "annual": 0,
                "vac": -120,
                "tags": "tag",
                "comment": "comment",
            },
            {
                "day": "mardi",
                "date": date(2026, 1, 2),
                "hr": 120,
                "annual": 30,
                "vac": 0,
                "tags": "",
                "comment": "",
            },
        ],
    )

    workbook = load_workbook(filename)
    worksheet = workbook["Records"]

    assert list(worksheet.iter_rows(values_only=True)) == [
        ("Day", "Date", "HR", "Annual", "VAC", "Tags", "Comment"),
        (
            "lundi",
            datetime(2026, 1, 1),
            timedelta(minutes=90),
            timedelta(0),
            timedelta(minutes=-120),
            "tag",
            "comment",
        ),
        (
            "mardi",
            datetime(2026, 1, 2),
            timedelta(minutes=120),
            timedelta(minutes=30),
            timedelta(0),
            None,
            None,
        ),
    ]
    assert worksheet["B2"].number_format == XlsxExporter.date_format
    assert worksheet["C2"].number_format == XlsxExporter.duration_format
    assert worksheet["D2"].number_format == XlsxExporter.duration_format
    assert worksheet["E2"].number_format == XlsxExporter.duration_format
